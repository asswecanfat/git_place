from openpyxl import *

wb = load_workbook('计科五班考勤表.xlsx')
ws = wb.active
li = ['C', 'D', 'E', 'F', 'G']
get = []
non = ''
for a in range(1,6):
    b = str(input('星期'+str(a)+'的情况:'))
    get.append(b)
    
for i in range(len(li)):
    ws[li[i]+'4'] = get[i]
    
for c in get:
    if c != '全勤':
        non = non + c + '  '

    
ws.cell(row=6, column=2, value=non)

wb.save('计科五班考勤表.xlsx')

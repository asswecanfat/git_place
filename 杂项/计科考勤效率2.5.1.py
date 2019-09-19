from openpyxl import *
from tkinter import *

def put_data():
    a = []
    for o in range(0, 5):
        a.append(e[o].get())
    a.append(r.get())
    get = a
    wb = load_workbook('计科五班考勤表.xlsx')
    ws = wb.active
    li = ['C', 'D', 'E', 'F', 'G']
    non = ''    
    for i in range(len(li)):
        ws[li[i]+'4'] = get[i]
        
    for c in range(5):
        if get == []:
            non = '全勤'
        elif get[c] != '':
            non = non + get[c] + '  '
        else :
            ws[li[c]+'4'] = '全勤'               
   
    ws.cell(row=6, column=2, value=non)

    wb.save(r'C:\Users\10248\Desktop\签到\第'+get[5]+'周.xlsx')
    for p in e:
        p.delete(0, END)
    print('成功输入')
    

tk  = Tk()
e = []
for i in range(0, 5):
    Label(tk, text='星期'+str(i+1)+'的情况:').grid(row = i,padx = 10, pady = 10)
    k = Entry(tk, width=100)
    k.grid(row=i, column = 1, padx = 10, pady = 10)
    e.append(k)
Label(tk, text = '周次:').grid(row = 5,padx = 10, pady = 10)
r = Entry(tk)
r.grid(row = 5, column = 1, sticky = W, padx = 10, pady = 10)
e.append(r)

Button(tk, text='导入数据', command=put_data).grid(row=6, column = 0, sticky = W, padx = 10, pady = 10)
Button(tk, text='退出', command=tk.quit).grid(row=6, column = 1, sticky = W, padx = 10, pady = 10)

mainloop()

